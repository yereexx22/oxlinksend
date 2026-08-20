import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk
from tkinter import Text as tkText
import re
import threading
import asyncio
from datetime import datetime
from pathlib import Path

import configuracion
import trabajadores
import plantillas
import mensajeria
from actualizador import Actualizador, VERSION_ACTUAL


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title('Oxlinksend - Gestor de Nóminas')
        self.geometry('1100x900')
        ctk.set_appearance_mode('System')
        ctk.set_default_color_theme('blue')

        self.config_correo = configuracion.cargar_config_correo()
        self.config_telegram = configuracion.cargar_config_telegram()

        self.tabview = ctk.CTkTabview(self)
        self.tabview.pack(fill='both', expand=True, padx=10, pady=10)

        self.tab_registro = self.tabview.add('Registrar Trabajadores')
        self.tab_envio = self.tabview.add('Enviar Nóminas (por ID)')
        self.tab_directo = self.tabview.add('Envío Directo')
        self.tab_reportes = self.tabview.add('Reportes')

        top_frame = ctk.CTkFrame(self, height=40)
        top_frame.pack(fill='x', padx=10, pady=(10, 0))
        self.btn_config_correo = ctk.CTkButton(top_frame, text='⚙ Configurar correo remitente', command=self.abrir_configuracion_correo)
        self.btn_config_correo.pack(side='right', padx=5)
        self.btn_config_telegram = ctk.CTkButton(top_frame, text='📱 Configurar Telegram', command=self.abrir_configuracion_telegram)
        self.btn_config_telegram.pack(side='right', padx=5)
        self.btn_plantilla = ctk.CTkButton(top_frame, text='✏ Editar plantilla de correo', command=self.editar_plantilla)
        self.btn_plantilla.pack(side='right', padx=5)
        self.btn_actualizar = ctk.CTkButton(top_frame, text='🔄 Buscar actualizaciones', command=self.buscar_actualizaciones, width=150)
        self.btn_actualizar.pack(side='right', padx=5)

        self.setup_tab_registro()
        self.setup_tab_envio()
        self.setup_tab_directo()
        self.setup_tab_reportes()
        self.iniciar_limpieza_automatica()

    # ------------------------------------------------------------
    # Métodos de actualización
    # ------------------------------------------------------------
    def buscar_actualizaciones(self):
        """Busca actualizaciones disponibles."""
        ventana = ctk.CTkToplevel(self)
        ventana.title("Verificando...")
        ventana.geometry("300x100")
        ventana.grab_set()
        
        ctk.CTkLabel(ventana, text="Buscando actualizaciones...").pack(pady=30)
        
        def verificar():
            actualizador = Actualizador()
            hay_actualizacion, version = actualizador.verificar_actualizacion()
            
            ventana.destroy()
            
            if hay_actualizacion:
                self.mostrar_ventana_actualizacion(actualizador, version)
            elif hay_actualizacion is False:
                messagebox.showinfo("Sin actualizaciones", "Ya tienes la última versión.")
            else:
                messagebox.showerror("Error", "No se pudo verificar la actualización.")
        
        threading.Thread(target=verificar, daemon=True).start()

    def mostrar_ventana_actualizacion(self, actualizador, version):
        """Muestra ventana de actualización disponible."""
        ventana = ctk.CTkToplevel(self)
        ventana.title("Actualización disponible")
        ventana.geometry("400x250")
        ventana.grab_set()
        
        frame = ctk.CTkFrame(ventana)
        frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        ctk.CTkLabel(
            frame,
            text="¡Nueva versión disponible! 🎉",
            font=ctk.CTkFont(size=18, weight='bold')
        ).pack(pady=15)
        
        ctk.CTkLabel(
            frame,
            text=f"Versión actual: {VERSION_ACTUAL}\nNueva versión: {version}",
            font=ctk.CTkFont(size=14)
        ).pack(pady=10)
        
        btn_frame = ctk.CTkFrame(frame)
        btn_frame.pack(pady=20)
        
        btn_actualizar = ctk.CTkButton(
            btn_frame,
            text="🔄 Actualizar ahora",
            command=lambda: self.iniciar_actualizacion(actualizador, ventana),
            fg_color='green',
            width=150
        )
        btn_actualizar.pack(side='left', padx=10)
        
        btn_cancelar = ctk.CTkButton(
            btn_frame,
            text="Cancelar",
            command=ventana.destroy,
            fg_color='gray',
            width=100
        )
        btn_cancelar.pack(side='left', padx=10)

    def iniciar_actualizacion(self, actualizador, ventana_info):
        """Inicia la actualización."""
        ventana_info.destroy()
        
        ventana = ctk.CTkToplevel(self)
        ventana.title("Actualizando...")
        ventana.geometry("400x150")
        ventana.grab_set()
        
        frame = ctk.CTkFrame(ventana)
        frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        label = ctk.CTkLabel(frame, text="Descargando...")
        label.pack(pady=10)
        
        barra = ctk.CTkProgressBar(frame, width=300)
        barra.pack(pady=10)
        barra.set(0)
        
        def progreso(texto, valor):
            label.configure(text=texto)
            barra.set(valor)
        
        def instalar():
            resultado = actualizador.descargar_e_instalar(progreso)
            ventana.destroy()
            if resultado:
                messagebox.showinfo("Éxito", "Actualización completada.\nReinicie la aplicación.")
            else:
                messagebox.showerror("Error", "No se pudo completar la actualización.")
        
        threading.Thread(target=instalar, daemon=True).start()

    # ------------------------------------------------------------
    # Tab: Registro de trabajadores
    # ------------------------------------------------------------
    def setup_tab_registro(self):
        form_frame = ctk.CTkFrame(self.tab_registro)
        form_frame.pack(fill='x', padx=10, pady=10)

        ctk.CTkLabel(form_frame, text='ID (número):').grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.id_input = ctk.CTkEntry(form_frame, width=150)
        self.id_input.grid(row=0, column=1, padx=5, pady=5)

        ctk.CTkLabel(form_frame, text='Nombre completo:').grid(row=1, column=0, padx=5, pady=5, sticky='e')
        self.nombre_input = ctk.CTkEntry(form_frame, width=250)
        self.nombre_input.grid(row=1, column=1, padx=5, pady=5)

        ctk.CTkLabel(form_frame, text='Correo electrónico:').grid(row=2, column=0, padx=5, pady=5, sticky='e')
        self.email_input = ctk.CTkEntry(form_frame, width=250)
        self.email_input.grid(row=2, column=1, padx=5, pady=5)

        ctk.CTkLabel(form_frame, text='Teléfono (ej: +584121598443):').grid(row=3, column=0, padx=5, pady=5, sticky='e')
        self.telefono_input = ctk.CTkEntry(form_frame, width=200)
        self.telefono_input.grid(row=3, column=1, padx=5, pady=5)

        btn_agregar = ctk.CTkButton(form_frame, text='Agregar Trabajador', command=self.agregar_trabajador)
        btn_agregar.grid(row=4, column=0, columnspan=2, pady=10)

        btn_importar_excel = ctk.CTkButton(form_frame, text='📥 Importar desde Excel', command=self.abrir_importar_excel)
        btn_importar_excel.grid(row=5, column=0, columnspan=2, pady=5)

        list_frame = ctk.CTkFrame(self.tab_registro)
        list_frame.pack(fill='both', expand=True, padx=10, pady=10)

        ctk.CTkLabel(list_frame, text='Trabajadores registrados', font=ctk.CTkFont(weight='bold')).pack(anchor='w', pady=(0, 5))

        search_frame = ctk.CTkFrame(list_frame)
        search_frame.pack(fill='x', pady=(0, 5))
        ctk.CTkLabel(search_frame, text='Buscar:').pack(side='left', padx=(5, 5))
        self.entry_buscar_registro = ctk.CTkEntry(search_frame, width=250)
        self.entry_buscar_registro.pack(side='left', padx=5)
        btn_buscar = ctk.CTkButton(search_frame, text='🔍 Buscar', command=self.buscar_trabajador_registro, width=100)
        btn_buscar.pack(side='left', padx=5)
        btn_limpiar = ctk.CTkButton(search_frame, text='Limpiar', command=self.limpiar_busqueda_registro, width=80, fg_color='gray')
        btn_limpiar.pack(side='left', padx=5)

        tree_frame = ctk.CTkFrame(list_frame)
        tree_frame.pack(fill='both', expand=True)

        scroll_y = ttk.Scrollbar(tree_frame, orient='vertical')
        self.tree = ttk.Treeview(tree_frame, columns=('ID', 'Nombre', 'Correo', 'Teléfono'), show='headings', yscrollcommand=scroll_y.set)
        scroll_y.config(command=self.tree.yview)
        scroll_y.pack(side='right', fill='y')
        self.tree.pack(side='left', fill='both', expand=True)

        self.tree.heading('ID', text='ID')
        self.tree.heading('Nombre', text='Nombre')
        self.tree.heading('Correo', text='Correo')
        self.tree.heading('Teléfono', text='Teléfono')
        self.tree.column('ID', width=50, anchor='center')
        self.tree.column('Nombre', width=180)
        self.tree.column('Correo', width=220)
        self.tree.column('Teléfono', width=120)

        btn_frame = ctk.CTkFrame(list_frame)
        btn_frame.pack(pady=10)
        btn_eliminar = ctk.CTkButton(btn_frame, text='Eliminar seleccionado', command=self.eliminar_trabajador, fg_color='red')
        btn_eliminar.pack(side='left', padx=5)
        btn_editar = ctk.CTkButton(btn_frame, text='✏ Editar seleccionado', command=self.editar_trabajador, fg_color='blue')
        btn_editar.pack(side='left', padx=5)

        self.refresh_tabla()

    def refresh_tabla(self, filtro=None):
        for item in self.tree.get_children():
            self.tree.delete(item)
        lista = trabajadores.listar_trabajadores()
        if filtro is not None:
            filtro = filtro.lower().strip()
            lista = [t for t in lista if filtro in str(t[0]) or filtro in t[1].lower() or filtro in t[2].lower()]
        for id_, nombre, email, telefono in lista:
            self.tree.insert('', 'end', values=(id_, nombre, email, telefono))

    def buscar_trabajador_registro(self):
        filtro = self.entry_buscar_registro.get().strip()
        if filtro:
            self.refresh_tabla(filtro)
        else:
            self.refresh_tabla()

    def limpiar_busqueda_registro(self):
        self.entry_buscar_registro.delete(0, 'end')
        self.refresh_tabla()

    def agregar_trabajador(self):
        try:
            id_ = int(self.id_input.get().strip())
        except ValueError:
            messagebox.showerror('Error', 'ID debe ser un número')
            return
        nombre = self.nombre_input.get().strip()
        email = self.email_input.get().strip()
        telefono = self.telefono_input.get().strip()
        if not nombre or not email:
            messagebox.showerror('Error', 'Nombre y correo son obligatorios')
            return

        if trabajadores.agregar_trabajador(id_, nombre, email, telefono):
            self.refresh_tabla()
            self.id_input.delete(0, 'end')
            self.nombre_input.delete(0, 'end')
            self.email_input.delete(0, 'end')
            self.telefono_input.delete(0, 'end')
            messagebox.showinfo('Éxito', f'Trabajador {nombre} agregado')
        else:
            messagebox.showerror('Error', f'Ya existe un trabajador con ID {id_}')

    def eliminar_trabajador(self):
        seleccion = self.tree.selection()
        if not seleccion:
            messagebox.showwarning('Advertencia', 'Seleccione un trabajador')
            return
        valores = self.tree.item(seleccion[0])['values']
        if valores:
            id_, nombre, _, _ = valores
            if messagebox.askyesno('Confirmar', f'¿Eliminar a {nombre} (ID {id_})?'):
                if trabajadores.eliminar_trabajador(id_):
                    self.refresh_tabla()
                    messagebox.showinfo('Éxito', 'Trabajador eliminado')

    def editar_trabajador(self):
        seleccion = self.tree.selection()
        if not seleccion:
            messagebox.showwarning('Advertencia', 'Seleccione un trabajador para editar')
            return
        valores = self.tree.item(seleccion[0])['values']
        if not valores:
            return
        id_actual, nombre_actual, email_actual, telefono_actual = valores

        dialog = ctk.CTkToplevel(self)
        dialog.title('Editar trabajador')
        dialog.geometry('500x350')
        dialog.resizable(False, False)
        dialog.grab_set()

        main_frame = ctk.CTkFrame(dialog)
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        main_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(main_frame, text='ID:').grid(row=0, column=0, padx=5, pady=10, sticky='e')
        id_entry = ctk.CTkEntry(main_frame, width=200)
        id_entry.insert(0, str(id_actual))
        id_entry.grid(row=0, column=1, padx=5, pady=10, sticky='w')

        ctk.CTkLabel(main_frame, text='Nombre:').grid(row=1, column=0, padx=5, pady=10, sticky='e')
        nombre_entry = ctk.CTkEntry(main_frame, width=200)
        nombre_entry.insert(0, nombre_actual)
        nombre_entry.grid(row=1, column=1, padx=5, pady=10, sticky='w')

        ctk.CTkLabel(main_frame, text='Correo:').grid(row=2, column=0, padx=5, pady=10, sticky='e')
        email_entry = ctk.CTkEntry(main_frame, width=200)
        email_entry.insert(0, email_actual)
        email_entry.grid(row=2, column=1, padx=5, pady=10, sticky='w')

        ctk.CTkLabel(main_frame, text='Teléfono:').grid(row=3, column=0, padx=5, pady=10, sticky='e')
        telefono_entry = ctk.CTkEntry(main_frame, width=200)
        telefono_entry.insert(0, telefono_actual)
        telefono_entry.grid(row=3, column=1, padx=5, pady=10, sticky='w')

        def guardar_edicion():
            try:
                nuevo_id = int(id_entry.get().strip())
            except ValueError:
                messagebox.showerror('Error', 'ID debe ser un número')
                return
            nuevo_nombre = nombre_entry.get().strip()
            nuevo_email = email_entry.get().strip()
            nuevo_telefono = telefono_entry.get().strip()
            if not nuevo_nombre or not nuevo_email:
                messagebox.showerror('Error', 'Nombre y correo son obligatorios')
                return
            if trabajadores.editar_trabajador(id_actual, nuevo_id, nuevo_nombre, nuevo_email, nuevo_telefono):
                self.refresh_tabla()
                dialog.destroy()
                messagebox.showinfo('Éxito', f'Trabajador {nuevo_nombre} actualizado')
            else:
                messagebox.showerror('Error', f'No se pudo editar. Verifica que el ID {nuevo_id} no esté en uso por otro trabajador.')

        btn_guardar = ctk.CTkButton(main_frame, text='Guardar cambios', command=guardar_edicion, fg_color='green')
        btn_guardar.grid(row=4, column=0, columnspan=2, pady=20)
        btn_cancelar = ctk.CTkButton(main_frame, text='Cancelar', command=dialog.destroy, fg_color='gray')
        btn_cancelar.grid(row=5, column=0, columnspan=2, pady=(0, 10))

    # ------------------------------------------------------------
    # Importación desde Excel
    # ------------------------------------------------------------
    def abrir_importar_excel(self):
        dialog = ctk.CTkToplevel(self)
        dialog.title('Importar trabajadores desde Excel')
        dialog.geometry('500x400')
        dialog.grab_set()

        main_frame = ctk.CTkFrame(dialog)
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)

        ctk.CTkLabel(main_frame, text='Importar desde archivo Excel', font=ctk.CTkFont(size=16, weight='bold')).pack(pady=10)

        self.ruta_excel_label = ctk.CTkLabel(main_frame, text='Ningún archivo seleccionado', fg_color='gray75', corner_radius=5)
        self.ruta_excel_label.pack(fill='x', pady=10)

        btn_sel_excel = ctk.CTkButton(main_frame, text='📂 Seleccionar archivo', command=self.seleccionar_excel)
        btn_sel_excel.pack(pady=5)

        btn_descargar_plantilla = ctk.CTkButton(main_frame, text='📄 Descargar plantilla', command=self.descargar_plantilla_excel)
        btn_descargar_plantilla.pack(pady=5)

        btn_importar = ctk.CTkButton(main_frame, text='🚀 Importar', command=self.importar_excel, fg_color='green')
        btn_importar.pack(pady=10)

        self.ruta_excel = None

    def seleccionar_excel(self):
        archivo = filedialog.askopenfilename(title='Seleccionar archivo Excel', filetypes=[('Excel files', '*.xlsx')])
        if archivo:
            self.ruta_excel = archivo
            self.ruta_excel_label.configure(text=Path(archivo).name)

    def descargar_plantilla_excel(self):
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Trabajadores'
            ws.append(['ID', 'Nombre', 'Email', 'Telefono'])
            ws.append([1, 'Ejemplo Nombre', 'ejemplo@correo.com', '+584121598443'])
            ruta_guardar = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='plantilla_trabajadores.xlsx'
            )
            if ruta_guardar:
                wb.save(ruta_guardar)
                messagebox.showinfo('Éxito', 'Plantilla descargada correctamente')
        except ImportError:
            messagebox.showerror('Error', 'Debe instalar openpyxl: pip install openpyxl')
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo guardar la plantilla: {e}')

    def importar_excel(self):
        if not self.ruta_excel:
            messagebox.showwarning('Advertencia', 'Seleccione un archivo Excel primero')
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.ruta_excel)
            ws = wb.active
            rows = ws.iter_rows(min_row=2, values_only=True)
            lista_importar = []
            errores_lectura = []
            for row in rows:
                if all(cell is None for cell in row):
                    continue
                if len(row) < 2:
                    errores_lectura.append(f'Fila incompleta: {row}')
                    continue
                id_ = row[0]
                nombre = row[1]
                email = row[2] if len(row) > 2 else ''
                telefono = row[3] if len(row) > 3 else ''
                try:
                    id_ = int(id_)
                except:
                    errores_lectura.append(f'ID inválido en fila: {row}')
                    continue
                if not nombre or not email:
                    errores_lectura.append(f'Fila con nombre o email vacío: {row}')
                    continue
                lista_importar.append((id_, str(nombre).strip(), str(email).strip(), str(telefono).strip()))
            if not lista_importar:
                messagebox.showerror('Error', 'No se encontraron filas válidas en el archivo')
                return
            agregados, errores_dup = trabajadores.importar_trabajadores(lista_importar)
            self.refresh_tabla()
            msg = f'Se agregaron {agregados} trabajadores.\n'
            if errores_dup:
                msg += 'Errores:\n' + '\n'.join(errores_dup) + '\n'
            if errores_lectura:
                msg += 'Advertencias de lectura:\n' + '\n'.join(errores_lectura)
            messagebox.showinfo('Resultado', msg)
        except ImportError:
            messagebox.showerror('Error', 'Debe instalar openpyxl: pip install openpyxl')
        except Exception as e:
            messagebox.showerror('Error', f'Error al leer el archivo: {e}')

    # ------------------------------------------------------------
    # Tab: Envío por lotes (por ID)
    # ------------------------------------------------------------
    def setup_tab_envio(self):
        instrucciones = ctk.CTkLabel(self.tab_envio, text='1. Seleccione uno o varios archivos PDF (nombre debe contener el ID del trabajador)\n2. La app buscará el ID en el nombre y enviará al correo/telegram registrado', justify='left', wraplength=800)
        instrucciones.pack(padx=10, pady=(10, 5), anchor='w')

        self.btn_seleccionar = ctk.CTkButton(self.tab_envio, text='📂 Seleccionar archivos PDF', command=self.seleccionar_pdfs)
        self.btn_seleccionar.pack(pady=5)

        list_group = ctk.CTkFrame(self.tab_envio)
        list_group.pack(fill='x', padx=10, pady=10)
        ctk.CTkLabel(list_group, text='Archivos seleccionados:', font=ctk.CTkFont(weight='bold')).pack(anchor='w')
        self.lista_archivos = tkText(list_group, height=5, state='normal', wrap='none')
        self.lista_archivos.pack(fill='both', expand=True, pady=5)
        self.lista_archivos.config(state='disabled')

        self.telegram_checkbox = ctk.CTkCheckBox(self.tab_envio, text='Enviar también por Telegram (pausa de 7-8.2 seg entre envíos)')
        self.telegram_checkbox.pack(pady=5)

        log_group = ctk.CTkFrame(self.tab_envio)
        log_group.pack(fill='both', expand=True, padx=10, pady=10)
        ctk.CTkLabel(log_group, text='Registro de envíos:', font=ctk.CTkFont(weight='bold')).pack(anchor='w')
        self.log_text = ctk.CTkTextbox(log_group, font=('Consolas', 9))
        self.log_text.pack(fill='both', expand=True, pady=5)
        self.log_text_widget = self.log_text._textbox
        self.log_text_widget.tag_config('success', foreground='green')
        self.log_text_widget.tag_config('error', foreground='red')
        self.log_text_widget.tag_config('normal', foreground='black')

        self.btn_enviar = ctk.CTkButton(self.tab_envio, text='🚀 Enviar nóminas seleccionadas', command=self.enviar_nominas, fg_color='green')
        self.btn_enviar.pack(pady=10)

        self.archivos_pdf = []

    def seleccionar_pdfs(self):
        archivos = filedialog.askopenfilenames(title='Seleccionar nóminas PDF', filetypes=[('PDF files', '*.pdf')])
        if archivos:
            self.archivos_pdf = list(archivos)
            self.lista_archivos.config(state='normal')
            self.lista_archivos.delete('1.0', 'end')
            for a in self.archivos_pdf:
                self.lista_archivos.insert('end', Path(a).name + '\n')
            self.lista_archivos.config(state='disabled')

    def enviar_nominas(self):
        if not self.archivos_pdf:
            messagebox.showwarning('Advertencia', 'No hay archivos seleccionados')
            return
        if not self.config_correo['EMAIL'] or not self.config_correo['PASSWORD']:
            messagebox.showerror('Error', 'Configure el correo remitente primero')
            return

        self.log_text.delete('1.0', 'end')
        self.append_log('Iniciando envíos...\n', 'normal')

        def log_callback(text, tipo):
            self.after(0, lambda: self.append_log(text, tipo))

        enviar_telegram = self.telegram_checkbox.get()
        hilos_correo = []
        cola_telegram = []

        for pdf in self.archivos_pdf:
            nombre_archivo = Path(pdf).name
            match = re.search(r'(\d+)', nombre_archivo)
            if not match:
                log_callback(f'⚠ No se encontró ID en {nombre_archivo}. Omitido.\n', 'error')
                continue
            id_ = int(match.group(1))
            trabajador = trabajadores.obtener_trabajador_por_id(id_)
            if not trabajador:
                log_callback(f'✗ No existe trabajador con ID {id_} (archivo: {nombre_archivo})\n', 'error')
                continue
            _, nombre, email, telefono = trabajador

            hilo_correo = threading.Thread(
                target=mensajeria.enviar_correo,
                args=(email, nombre, id_, pdf, log_callback, self.config_correo)
            )
            hilo_correo.start()
            hilos_correo.append(hilo_correo)

            if enviar_telegram and telefono:
                cola_telegram.append((telefono, nombre, id_, pdf, log_callback))
            elif enviar_telegram and not telefono:
                log_callback(f'⚠ No se envió por Telegram a {nombre} porque no tiene número registrado.\n', 'error')

        def procesar_telegram():
            for telefono, nombre, id_, pdf, log_cb in cola_telegram:
                mensajeria.enviar_telegram_sincrono(telefono, nombre, id_, pdf, log_cb)
            self.after(0, lambda: self.append_log('\n✅ Proceso de Telegram completado.\n', 'success'))

        def esperar_correos():
            for h in hilos_correo:
                h.join()
            if cola_telegram:
                threading.Thread(target=procesar_telegram).start()
            else:
                self.after(0, lambda: self.append_log('\n✅ Proceso completado.\n', 'success'))

        threading.Thread(target=esperar_correos).start()

    def append_log(self, texto, tipo):
        self.log_text_widget.insert('end', texto)
        end_index = self.log_text_widget.index('end-1c')
        start_index = f'{end_index} - {len(texto)}c'
        self.log_text_widget.tag_add(tipo, start_index, end_index)
        self.log_text_widget.see('end')

    # ------------------------------------------------------------
    # Tab: Envío directo
    # ------------------------------------------------------------
    def setup_tab_directo(self):
        container = ctk.CTkFrame(self.tab_directo)
        container.pack(fill='both', expand=True)

        bottom_frame = ctk.CTkFrame(container, height=60)
        bottom_frame.pack(side='bottom', fill='x', padx=10, pady=(0, 10))
        self.btn_enviar_directo = ctk.CTkButton(bottom_frame, text='📧 Enviar PDF directamente', command=self.enviar_directo, fg_color='green', height=40, font=ctk.CTkFont(size=14, weight='bold'))
        self.btn_enviar_directo.pack(fill='x', padx=20, pady=10)

        scrollable = ctk.CTkScrollableFrame(container, label_text='')
        scrollable.pack(side='top', fill='both', expand=True, padx=10, pady=(10, 5))

        ctk.CTkLabel(scrollable, text='Seleccionar trabajador (opcional):', font=ctk.CTkFont(weight='bold')).pack(anchor='w', pady=(0, 5))

        search_frame = ctk.CTkFrame(scrollable)
        search_frame.pack(fill='x', pady=5)
        ctk.CTkLabel(search_frame, text='Buscar:').pack(side='left', padx=(5, 5))
        self.entry_buscar_directo = ctk.CTkEntry(search_frame, width=250)
        self.entry_buscar_directo.pack(side='left', padx=5)
        self.entry_buscar_directo.bind('<KeyRelease>', self.filtrar_workers_directo)

        worker_frame = ctk.CTkFrame(scrollable)
        worker_frame.pack(fill='x', pady=5)
        self.worker_combo = ctk.CTkComboBox(worker_frame, values=[''], state='readonly', width=300)
        self.worker_combo.pack(side='left', padx=(0, 10))
        self.btn_refresh = ctk.CTkButton(worker_frame, text='⟳ Actualizar lista', command=self.actualizar_lista_workers, width=120)
        self.btn_refresh.pack(side='left')

        ctk.CTkLabel(scrollable, text='O ingresar datos manualmente:', font=ctk.CTkFont(weight='bold')).pack(anchor='w', pady=(15, 5))
        manual_frame = ctk.CTkFrame(scrollable)
        manual_frame.pack(fill='x', pady=5)

        ctk.CTkLabel(manual_frame, text='Nombre:').grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.nombre_manual = ctk.CTkEntry(manual_frame, width=250)
        self.nombre_manual.grid(row=0, column=1, padx=5, pady=5)

        ctk.CTkLabel(manual_frame, text='Correo:').grid(row=1, column=0, padx=5, pady=5, sticky='e')
        self.email_manual = ctk.CTkEntry(manual_frame, width=250)
        self.email_manual.grid(row=1, column=1, padx=5, pady=5)

        ctk.CTkLabel(manual_frame, text='Teléfono (ej: +584121598443):').grid(row=2, column=0, padx=5, pady=5, sticky='e')
        self.telefono_manual = ctk.CTkEntry(manual_frame, width=200)
        self.telefono_manual.grid(row=2, column=1, padx=5, pady=5)

        ctk.CTkLabel(scrollable, text='Seleccionar archivo PDF:', font=ctk.CTkFont(weight='bold')).pack(anchor='w', pady=(15, 5))
        pdf_frame = ctk.CTkFrame(scrollable)
        pdf_frame.pack(fill='x', pady=5)
        self.ruta_pdf_label = ctk.CTkLabel(pdf_frame, text='Ningún archivo seleccionado', fg_color='gray75', corner_radius=5, anchor='w')
        self.ruta_pdf_label.pack(side='left', padx=(0, 10), fill='x', expand=True)
        self.btn_pdf = ctk.CTkButton(pdf_frame, text='📄 Seleccionar PDF', command=self.seleccionar_pdf_directo)
        self.btn_pdf.pack(side='right')
        self.pdf_seleccionado = None

        self.telegram_directo_checkbox = ctk.CTkCheckBox(scrollable, text='Enviar también por Telegram (pausa de 7-8.2 seg)')
        self.telegram_directo_checkbox.pack(anchor='w', pady=10)

        log_frame = ctk.CTkFrame(scrollable)
        log_frame.pack(fill='x', pady=(20, 10))
        ctk.CTkLabel(log_frame, text='Registro de envío:', font=ctk.CTkFont(weight='bold')).pack(anchor='w')
        self.log_directo = ctk.CTkTextbox(log_frame, font=('Consolas', 9), height=150)
        self.log_directo.pack(fill='both', expand=True, pady=5)
        self.log_directo_widget = self.log_directo._textbox
        self.log_directo_widget.tag_config('success', foreground='green')
        self.log_directo_widget.tag_config('error', foreground='red')
        self.log_directo_widget.tag_config('normal', foreground='black')

        self.actualizar_lista_workers()
        self.worker_combo.bind('<<ComboboxSelected>>', self.on_worker_selected)

    def actualizar_lista_workers(self):
        self.todos_workers = trabajadores.listar_trabajadores()
        self.filtrar_workers_directo(reset=True)

    def filtrar_workers_directo(self, event=None, reset=False):
        if reset:
            filtro = ''
        else:
            filtro = self.entry_buscar_directo.get().strip().lower()
        lista = self.todos_workers
        if filtro:
            lista = [t for t in lista if filtro in str(t[0]) or filtro in t[1].lower() or filtro in t[2].lower()]
        nombres = [f'{id_} - {nombre}' for id_, nombre, _, _ in lista]
        if nombres:
            self.worker_combo.configure(values=nombres)
            if self.worker_combo.get() not in nombres:
                self.worker_combo.set(nombres[0])
                self.on_worker_selected()
        else:
            self.worker_combo.configure(values=['(Sin resultados)'])
            self.worker_combo.set('(Sin resultados)')

    def on_worker_selected(self, event=None):
        seleccion = self.worker_combo.get()
        if seleccion and seleccion != '(Sin resultados)' and seleccion != '(No hay trabajadores)':
            id_str = seleccion.split(' - ')[0]
            try:
                trabajador = trabajadores.obtener_trabajador_por_id(int(id_str))
                if trabajador:
                    _, nombre, email, telefono = trabajador
                    self.nombre_manual.delete(0, 'end')
                    self.nombre_manual.insert(0, nombre)
                    self.email_manual.delete(0, 'end')
                    self.email_manual.insert(0, email)
                    self.telefono_manual.delete(0, 'end')
                    self.telefono_manual.insert(0, telefono)
            except:
                pass

    def seleccionar_pdf_directo(self):
        archivo = filedialog.askopenfilename(title='Seleccionar PDF', filetypes=[('PDF files', '*.pdf')])
        if archivo:
            self.pdf_seleccionado = archivo
            self.ruta_pdf_label.configure(text=Path(archivo).name)

    def append_log_directo(self, texto, tipo):
        self.log_directo_widget.insert('end', texto)
        end_index = self.log_directo_widget.index('end-1c')
        start_index = f'{end_index} - {len(texto)}c'
        self.log_directo_widget.tag_add(tipo, start_index, end_index)
        self.log_directo_widget.see('end')

    def enviar_directo(self):
        nombre = self.nombre_manual.get().strip()
        email = self.email_manual.get().strip()
        telefono = self.telefono_manual.get().strip()
        seleccion = self.worker_combo.get()
        id_trabajador = None

        if seleccion and seleccion not in ['(No hay trabajadores)', '(Sin resultados)'] and not (nombre and email):
            try:
                id_str = seleccion.split(' - ')[0]
                id_trabajador = int(id_str)
                trabajador = trabajadores.obtener_trabajador_por_id(id_trabajador)
                if trabajador:
                    _, nombre_auto, email_auto, telefono_auto = trabajador
                    if not nombre:
                        nombre = nombre_auto
                        self.nombre_manual.delete(0, 'end')
                        self.nombre_manual.insert(0, nombre)
                    if not email:
                        email = email_auto
                        self.email_manual.delete(0, 'end')
                        self.email_manual.insert(0, email)
                    if not telefono:
                        telefono = telefono_auto
                        self.telefono_manual.delete(0, 'end')
                        self.telefono_manual.insert(0, telefono)
            except:
                pass

        if not self.pdf_seleccionado:
            messagebox.showwarning('Advertencia', 'No ha seleccionado ningún archivo PDF')
            return
        if not self.config_correo['EMAIL'] or not self.config_correo['PASSWORD']:
            messagebox.showerror('Error', 'Configure el correo remitente primero')
            return
        if not nombre or not email:
            messagebox.showerror('Error', 'Debe proporcionar el nombre y correo del destinatario')
            return

        self.log_directo.delete('1.0', 'end')
        self.append_log_directo('Enviando...\n', 'normal')

        def log_callback(text, tipo):
            self.after(0, lambda: self.append_log_directo(text, tipo))

        id_envio = id_trabajador if id_trabajador else 0
        hilo_correo = threading.Thread(target=mensajeria.enviar_correo, args=(email, nombre, id_envio, self.pdf_seleccionado, log_callback, self.config_correo))
        hilo_correo.start()

        enviar_telegram = self.telegram_directo_checkbox.get()
        hilo_telegram = None
        if enviar_telegram and telefono:
            hilo_telegram = threading.Thread(target=mensajeria.enviar_telegram_sincrono, args=(telefono, nombre, id_envio, self.pdf_seleccionado, log_callback))
            hilo_telegram.start()
        elif enviar_telegram and not telefono:
            self.append_log_directo('⚠ No se enviará por Telegram porque no se proporcionó número.\n', 'error')

        def esperar():
            hilo_correo.join()
            if hilo_telegram:
                hilo_telegram.join()
            self.after(0, lambda: self.append_log_directo('\n✅ Envío finalizado.\n', 'success'))

        threading.Thread(target=esperar).start()

    # ------------------------------------------------------------
    # Tab: Reportes
    # ------------------------------------------------------------
    def setup_tab_reportes(self):
        container = ctk.CTkFrame(self.tab_reportes)
        container.pack(fill='both', expand=True)

        btn_frame = ctk.CTkFrame(container)
        btn_frame.pack(fill='x', padx=10, pady=10)

        ctk.CTkLabel(btn_frame, text='Reporte de envíos exitosos', font=ctk.CTkFont(size=16, weight='bold')).pack(side='left', padx=5)

        btn_exportar = ctk.CTkButton(btn_frame, text='📥 Exportar a Excel', command=self.exportar_reportes_excel)
        btn_exportar.pack(side='right', padx=5)

        btn_actualizar = ctk.CTkButton(btn_frame, text='🔄 Actualizar', command=self.cargar_reportes)
        btn_actualizar.pack(side='right', padx=5)

        btn_eliminar = ctk.CTkButton(btn_frame, text='🗑 Eliminar seleccionado', command=self.eliminar_reporte, fg_color='red')
        btn_eliminar.pack(side='right', padx=5)

        tree_frame = ctk.CTkFrame(container)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=(0, 10))

        scroll_y = ttk.Scrollbar(tree_frame, orient='vertical')
        self.tree_reportes = ttk.Treeview(
            tree_frame,
            columns=('Fecha', 'ID', 'Nombre', 'Método', 'Archivo'),
            show='headings',
            yscrollcommand=scroll_y.set
        )
        scroll_y.config(command=self.tree_reportes.yview)
        scroll_y.pack(side='right', fill='y')
        self.tree_reportes.pack(side='left', fill='both', expand=True)

        self.tree_reportes.heading('Fecha', text='Fecha', command=lambda: self.ordenar_reportes('Fecha'))
        self.tree_reportes.heading('ID', text='ID', command=lambda: self.ordenar_reportes('ID'))
        self.tree_reportes.heading('Nombre', text='Nombre', command=lambda: self.ordenar_reportes('Nombre'))
        self.tree_reportes.heading('Método', text='Método', command=lambda: self.ordenar_reportes('Método'))
        self.tree_reportes.heading('Archivo', text='Archivo', command=lambda: self.ordenar_reportes('Archivo'))

        self.tree_reportes.column('Fecha', width=150)
        self.tree_reportes.column('ID', width=50, anchor='center')
        self.tree_reportes.column('Nombre', width=180)
        self.tree_reportes.column('Método', width=100, anchor='center')
        self.tree_reportes.column('Archivo', width=200)

        self.reportes_datos = []
        self.reportes_orden_actual = {}

        self.cargar_reportes()

    def cargar_reportes(self):
        for item in self.tree_reportes.get_children():
            self.tree_reportes.delete(item)

        self.reportes_datos = configuracion.cargar_envios_exitosos()
        self._insertar_reportes_en_tabla(self.reportes_datos)

    def _insertar_reportes_en_tabla(self, datos):
        for id_envio, fecha, id_trabajador, nombre, metodo, archivo in datos:
            item = self.tree_reportes.insert('', 'end', values=(fecha, id_trabajador, nombre, metodo, archivo))
            self.tree_reportes.item(item, tags=(str(id_envio),))

    def ordenar_reportes(self, columna):
        if columna not in self.reportes_orden_actual:
            self.reportes_orden_actual[columna] = 'asc'
        else:
            if self.reportes_orden_actual[columna] == 'asc':
                self.reportes_orden_actual[columna] = 'desc'
            else:
                self.reportes_orden_actual[columna] = 'asc'

        direccion = self.reportes_orden_actual[columna]

        indices = {
            'Fecha': 1,
            'ID': 2,
            'Nombre': 3,
            'Método': 4,
            'Archivo': 5,
        }
        idx = indices[columna]

        datos_ordenados = sorted(
            self.reportes_datos,
            key=lambda x: x[idx],
            reverse=(direccion == 'desc')
        )

        for item in self.tree_reportes.get_children():
            self.tree_reportes.delete(item)
        self._insertar_reportes_en_tabla(datos_ordenados)

    def eliminar_reporte(self):
        seleccion = self.tree_reportes.selection()
        if not seleccion:
            messagebox.showwarning('Advertencia', 'Seleccione un reporte para eliminar')
            return

        item = seleccion[0]
        tags = self.tree_reportes.item(item, 'tags')
        if not tags:
            return
        id_envio = int(tags[0])

        if not messagebox.askyesno('Confirmar', f'¿Eliminar el reporte con ID interno {id_envio}?'):
            return

        if configuracion.eliminar_envio(id_envio):
            self.cargar_reportes()
            messagebox.showinfo('Éxito', 'Reporte eliminado')
        else:
            messagebox.showerror('Error', 'No se pudo eliminar el reporte')

    def exportar_reportes_excel(self):
        items = self.tree_reportes.get_children()
        if not items:
            messagebox.showwarning('Advertencia', 'No hay reportes para exportar')
            return
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Reportes'
            ws.append(['Fecha', 'ID', 'Nombre', 'Método', 'Archivo'])
            for item in items:
                valores = self.tree_reportes.item(item)['values']
                ws.append(valores)
            ruta_guardar = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='reportes_envios.xlsx'
            )
            if ruta_guardar:
                wb.save(ruta_guardar)
                messagebox.showinfo('Éxito', 'Reporte exportado correctamente')
        except ImportError:
            messagebox.showerror('Error', 'Debe instalar openpyxl: pip install openpyxl')
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo exportar: {e}')

    # ------------------------------------------------------------
    # Limpieza automática de reportes
    # ------------------------------------------------------------
    def limpiar_reportes_automatico(self):
        hoy = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        fecha_limite = hoy.isoformat()
        configuracion.eliminar_envios_anteriores_a(fecha_limite)
        self.cargar_reportes()

    def iniciar_limpieza_automatica(self):
        self.limpiar_reportes_automatico()
        self.after(24 * 60 * 60 * 1000, self.iniciar_limpieza_automatica)

    # ------------------------------------------------------------
    # Diálogos de configuración
    # ------------------------------------------------------------
    def abrir_configuracion_correo(self):
        dialog = ctk.CTkToplevel(self)
        dialog.title('Configuración de correo remitente')
        dialog.geometry('500x300')
        dialog.resizable(True, True)
        dialog.grab_set()

        main_frame = ctk.CTkFrame(dialog)
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        main_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(main_frame, text='Correo Gmail:').grid(row=0, column=0, padx=5, pady=10, sticky='e')
        email_entry = ctk.CTkEntry(main_frame)
        email_entry.grid(row=0, column=1, padx=5, pady=10, sticky='ew')

        ctk.CTkLabel(main_frame, text='Contraseña (App Password):').grid(row=1, column=0, padx=5, pady=10, sticky='e')
        pass_entry = ctk.CTkEntry(main_frame, show='*')
        pass_entry.grid(row=1, column=1, padx=5, pady=10, sticky='ew')

        nota = ctk.CTkLabel(main_frame, text='Nota: Para Gmail necesitas una \'Contraseña de aplicación\'', text_color='gray')
        nota.grid(row=2, column=0, columnspan=2, pady=10)

        def guardar():
            email = email_entry.get().strip()
            password = pass_entry.get().strip()
            if email and password:
                configuracion.guardar_config_correo(email, password)
                self.config_correo = configuracion.cargar_config_correo()
                messagebox.showinfo('Éxito', 'Configuración guardada')
                dialog.destroy()
            else:
                messagebox.showwarning('Advertencia', 'Complete ambos campos')

        btn_guardar = ctk.CTkButton(main_frame, text='Guardar', command=guardar)
        btn_guardar.grid(row=3, column=0, columnspan=2, pady=20)

        config = configuracion.cargar_config_correo()
        email_entry.insert(0, config['EMAIL'])
        pass_entry.insert(0, config['PASSWORD'])

    def abrir_configuracion_telegram(self):
        dialog = ctk.CTkToplevel(self)
        dialog.title('Configuración de Telegram (Userbot)')
        dialog.geometry('600x500')
        dialog.resizable(True, True)
        dialog.grab_set()

        main_frame = ctk.CTkFrame(dialog)
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        main_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(main_frame, text='API ID:').grid(row=0, column=0, padx=5, pady=10, sticky='e')
        api_id_entry = ctk.CTkEntry(main_frame)
        api_id_entry.grid(row=0, column=1, padx=5, pady=10, sticky='ew')

        ctk.CTkLabel(main_frame, text='API Hash:').grid(row=1, column=0, padx=5, pady=10, sticky='e')
        api_hash_entry = ctk.CTkEntry(main_frame)
        api_hash_entry.grid(row=1, column=1, padx=5, pady=10, sticky='ew')

        ctk.CTkLabel(main_frame, text='Tu número de teléfono (ej: +584121598443):').grid(row=2, column=0, padx=5, pady=10, sticky='e')
        telefono_entry = ctk.CTkEntry(main_frame)
        telefono_entry.grid(row=2, column=1, padx=5, pady=10, sticky='ew')

        def guardar_config():
            api_id = api_id_entry.get().strip()
            api_hash = api_hash_entry.get().strip()
            telefono = telefono_entry.get().strip()
            if api_id and api_hash and telefono:
                configuracion.guardar_config_telegram(api_id, api_hash, telefono)
                self.config_telegram = configuracion.cargar_config_telegram()
                messagebox.showinfo('Éxito', 'Configuración de Telegram guardada')
            else:
                messagebox.showwarning('Advertencia', 'Complete todos los campos')

        btn_guardar = ctk.CTkButton(main_frame, text='Guardar configuración', command=guardar_config)
        btn_guardar.grid(row=3, column=0, columnspan=2, pady=10)

        codigo_event = threading.Event()
        codigo_ingresado = None
        password_event = threading.Event()
        password_ingresado = None

        def mostrar_dialogo_codigo():
            nonlocal codigo_ingresado
            code_dialog = ctk.CTkToplevel(dialog)
            code_dialog.title('Código de verificación')
            code_dialog.geometry('400x200')
            code_dialog.grab_set()
            ctk.CTkLabel(code_dialog, text='Se ha enviado un código a tu Telegram.\nIngrésalo aquí:').pack(pady=20)
            entry = ctk.CTkEntry(code_dialog, width=200)
            entry.pack(pady=10)

            def aceptar():
                nonlocal codigo_ingresado
                codigo_ingresado = entry.get().strip()
                code_dialog.destroy()
                codigo_event.set()

            btn_ok = ctk.CTkButton(code_dialog, text='Aceptar', command=aceptar)
            btn_ok.pack(pady=10)
            code_dialog.bind('<Return>', lambda e: aceptar())
            code_dialog.protocol('WM_DELETE_WINDOW', lambda: (code_dialog.destroy(), codigo_event.set()))

        def mostrar_dialogo_password():
            nonlocal password_ingresado
            pwd_dialog = ctk.CTkToplevel(dialog)
            pwd_dialog.title('Contraseña de 2FA')
            pwd_dialog.geometry('400x200')
            pwd_dialog.grab_set()
            ctk.CTkLabel(pwd_dialog, text='Ingresa tu contraseña de verificación en dos pasos:').pack(pady=20)
            entry = ctk.CTkEntry(pwd_dialog, width=200, show='*')
            entry.pack(pady=10)

            def aceptar():
                nonlocal password_ingresado
                password_ingresado = entry.get().strip()
                pwd_dialog.destroy()
                password_event.set()

            btn_ok = ctk.CTkButton(pwd_dialog, text='Aceptar', command=aceptar)
            btn_ok.pack(pady=10)
            pwd_dialog.bind('<Return>', lambda e: aceptar())
            pwd_dialog.protocol('WM_DELETE_WINDOW', lambda: (pwd_dialog.destroy(), password_event.set()))

        def autenticar():
            config = configuracion.cargar_config_telegram()
            if not config['api_id'] or not config['api_hash'] or not config['telefono_remitente']:
                messagebox.showerror('Error', 'Primero guarda la configuración de Telegram')
                return

            progress = ctk.CTkToplevel(dialog)
            progress.title('Autenticando...')
            progress.geometry('300x120')
            progress.grab_set()
            ctk.CTkLabel(progress, text='Conectando con Telegram...').pack(pady=20)
            progress.update()

            def auth_task():
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                try:
                    async def auth():
                        from telethon import TelegramClient
                        from telethon.sessions import StringSession
                        client = TelegramClient(
                            StringSession(),
                            int(config['api_id']),
                            config['api_hash']
                        )
                        await client.connect()
                        if not await client.is_user_authorized():
                            await client.send_code_request(config['telefono_remitente'])
                            self.after(0, mostrar_dialogo_codigo)
                            while not codigo_event.is_set():
                                await asyncio.sleep(0.1)
                            code = codigo_ingresado
                            if not code:
                                raise Exception('No se ingresó el código')
                            try:
                                await client.sign_in(config['telefono_remitente'], code)
                            except Exception as e:
                                if 'password' in str(e).lower():
                                    self.after(0, mostrar_dialogo_password)
                                    while not password_event.is_set():
                                        await asyncio.sleep(0.1)
                                    pwd = password_ingresado
                                    if pwd:
                                        await client.sign_in(password=pwd)
                                    else:
                                        raise Exception('No se ingresó la contraseña de 2FA')
                                else:
                                    raise
                        session_string = client.session.save()
                        configuracion.guardar_session_telegram_local(session_string)
                        await client.disconnect()
                        return True
                    loop.run_until_complete(auth())
                    self.after(0, progress.destroy)
                    self.after(0, lambda: messagebox.showinfo('Éxito', 'Autenticación de Telegram completada exitosamente'))
                except Exception as e:
                    self.after(0, progress.destroy)
                    self.after(0, lambda: messagebox.showerror('Error', f'Error de autenticación: {str(e)}'))
                finally:
                    loop.close()

            threading.Thread(target=auth_task, daemon=True).start()

        btn_autenticar = ctk.CTkButton(main_frame, text='🔐 Autenticar (iniciar sesión)', command=autenticar, fg_color='green')
        btn_autenticar.grid(row=4, column=0, columnspan=2, pady=10)

        nota = ctk.CTkLabel(main_frame, text='Nota: Necesitas un api_id y api_hash de my.telegram.org\nEl userbot puede violar términos de servicio, úsalo bajo tu responsabilidad.', text_color='orange', wraplength=500)
        nota.grid(row=5, column=0, columnspan=2, pady=10)

        config = configuracion.cargar_config_telegram()
        api_id_entry.insert(0, config['api_id'])
        api_hash_entry.insert(0, config['api_hash'])
        telefono_entry.insert(0, config['telefono_remitente'])

    def editar_plantilla(self):
        dialog = ctk.CTkToplevel(self)
        dialog.title('Editar plantilla de correo')
        dialog.geometry('650x550')
        dialog.resizable(True, True)
        dialog.grab_set()

        main_frame = ctk.CTkFrame(dialog)
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        main_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(main_frame, text='Asunto:').grid(row=0, column=0, sticky='w', pady=(0, 5))
        asunto_entry = ctk.CTkEntry(main_frame)
        asunto_entry.grid(row=1, column=0, sticky='ew', pady=(0, 15))

        ctk.CTkLabel(main_frame, text='Cuerpo:').grid(row=2, column=0, sticky='w', pady=(0, 5))
        cuerpo_text = ctk.CTkTextbox(main_frame, height=300)
        cuerpo_text.grid(row=3, column=0, sticky='nsew', pady=(0, 10))
        main_frame.grid_rowconfigure(3, weight=1)

        info = ctk.CTkLabel(main_frame, text='Usa {nombre} para insertar el nombre del trabajador', text_color='gray')
        info.grid(row=4, column=0, sticky='w', pady=(0, 10))

        asunto_actual, cuerpo_actual = plantillas.cargar_plantilla()
        asunto_entry.insert(0, asunto_actual)
        cuerpo_text.insert('1.0', cuerpo_actual)

        def guardar():
            nuevo_asunto = asunto_entry.get().strip()
            nuevo_cuerpo = cuerpo_text.get('1.0', 'end-1c').strip()
            if not nuevo_asunto or not nuevo_cuerpo:
                messagebox.showwarning('Advertencia', 'Asunto y cuerpo no pueden estar vacíos')
                return
            plantillas.guardar_plantilla(nuevo_asunto, nuevo_cuerpo)
            messagebox.showinfo('Éxito', 'Plantilla guardada')
            dialog.destroy()

        btn_guardar = ctk.CTkButton(main_frame, text='Guardar plantilla', command=guardar)
        btn_guardar.grid(row=5, column=0, pady=10)

    def on_closing(self):
        self.destroy()